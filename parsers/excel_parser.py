from __future__ import annotations

import csv
import os

from .base import ParsedDocument, ParsedPage
from utils.logger import get_logger

logger = get_logger(__name__)


def parse_excel(file_path: str) -> ParsedDocument:
    """Extract text and tables from Excel/CSV files."""
    ext = os.path.splitext(file_path)[1].lower()

    if ext == ".csv":
        return _parse_csv(file_path)
    return _parse_xlsx(file_path)


def _parse_xlsx(file_path: str) -> ParsedDocument:
    from openpyxl import load_workbook

    wb = load_workbook(file_path, read_only=True, data_only=True)
    pages: list[ParsedPage] = []

    for sheet_idx, sheet_name in enumerate(wb.sheetnames):
        ws = wb[sheet_name]
        rows: list[list[str]] = []
        text_parts: list[str] = []

        for row in ws.iter_rows(values_only=True):
            cells = [str(cell) if cell is not None else "" for cell in row]
            # Skip completely empty rows
            if any(c.strip() for c in cells):
                rows.append(cells)

        # Build readable text from table rows
        if rows:
            # First row as header
            header = rows[0]
            text_parts.append(f"[工作表: {sheet_name}]")
            text_parts.append(" | ".join(header))
            text_parts.append("-" * 40)
            for row in rows[1:]:
                # Build key-value style text for better semantic chunking
                kv_pairs = []
                for h, v in zip(header, row):
                    if v.strip() and h.strip():
                        kv_pairs.append(f"{h}: {v}")
                if kv_pairs:
                    text_parts.append("; ".join(kv_pairs))

        pages.append(
            ParsedPage(
                page_number=sheet_idx + 1,
                text="\n".join(text_parts),
                tables=[rows] if rows else [],
            )
        )

    wb.close()
    logger.info(f"Excel parsed: {file_path} ({len(pages)} sheets)")
    return ParsedDocument(
        file_name=os.path.basename(file_path),
        file_path=file_path,
        file_type="excel",
        pages=pages,
        metadata={"total_sheets": len(pages), "sheet_names": list(wb.sheetnames) if hasattr(wb, 'sheetnames') else []},
    )


def _parse_csv(file_path: str) -> ParsedDocument:
    rows: list[list[str]] = []
    text_parts: list[str] = []

    # Try to detect encoding
    encoding = "utf-8"
    try:
        with open(file_path, "r", encoding="utf-8") as f:
            f.read(1024)
    except UnicodeDecodeError:
        encoding = "gbk"

    with open(file_path, "r", encoding=encoding, errors="replace") as f:
        reader = csv.reader(f)
        for row in reader:
            cells = [cell.strip() for cell in row]
            if any(cells):
                rows.append(cells)

    if rows:
        header = rows[0]
        text_parts.append(" | ".join(header))
        text_parts.append("-" * 40)
        for row in rows[1:]:
            kv_pairs = []
            for h, v in zip(header, row):
                if v.strip() and h.strip():
                    kv_pairs.append(f"{h}: {v}")
            if kv_pairs:
                text_parts.append("; ".join(kv_pairs))

    page = ParsedPage(
        page_number=1,
        text="\n".join(text_parts),
        tables=[rows] if rows else [],
    )

    logger.info(f"CSV parsed: {file_path} ({len(rows)} rows)")
    return ParsedDocument(
        file_name=os.path.basename(file_path),
        file_path=file_path,
        file_type="excel",
        pages=[page],
        metadata={"total_rows": len(rows)},
    )
